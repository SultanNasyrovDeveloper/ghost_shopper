import os
from html.parser import HTMLParser
from pathlib import Path

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.writer.excel import save_virtual_workbook

from ghost_shopper.check.models import Check
from ghost_shopper.organisation_tree.statistics import ChecksStatistics


class Parser(HTMLParser):
    """Parser for question text strings"""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._data = ''

    @property
    def data(self):
        return self._data.strip(', ')

    def handle_data(self, data):
        self._data += data + ' '

    def clean(self):
        self._data = ''


class ChecklistToExcelConverter:
    """ Converting checklist data into excel file """
    def __init__(self, checklist):
        self.checklist = checklist
        self.file_name = self._generate_name()
        self.workbook = Workbook()
        self.ws = self.workbook.active
        qs = Check.objects.filter(id=self.checklist.check_obj_id)
        self.statistics = ChecksStatistics(qs)
        self.statistics.calculate()
        self.parser = Parser()
        self.styles = {
            'bold_font': Font(bold=True),
            'centered_text': Alignment(horizontal='center', wrap_text=True, vertical='center'),
            'wrap_text': Alignment(wrap_text=True, vertical='center'),
            'border_1': Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                bottom=Side(border_style='thin'),
                top=Side(border_style='thin'),
            ),
            'border_2': Border(
                left=Side(border_style='thick'),
                right=Side(border_style='thick'),
                bottom=Side(border_style='thick'),
                top=Side(border_style='thick'),
            ),
            'bg_grey': PatternFill(
                start_color='D6CFC7',
                end_color='D6CFC7',
                fill_type='solid'
            ),
            'question_font': Font(size=10),
            'question_align': Alignment(wrap_text=True, vertical='center'),
            'question_cell_border': Border(
                left=Side(border_style='thick'),
                right=Side(border_style='thin'),
                bottom=Side(border_style='thin'),
            ),
            'usual_border': Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                bottom=Side(border_style='thin'),
            ),
            'usual_border_right': Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thick'),
                bottom=Side(border_style='thin'),
            ),
            'botton_thick_border': Border(
                bottom=Side(border_style='thick'),
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
            ),
            'last_row_left': Border(
                bottom=Side(border_style='thick'),
                left=Side(border_style='thick'),
                right=Side(border_style='thin'),
            ),
            'last_row_center': Border(
                bottom=Side(border_style='thick'),
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
            ),
            'last_row_right': Border(
                bottom=Side(border_style='thick'),
                left=Side(border_style='thin'),
                right=Side(border_style='thick'),
            ),
        }

    @property
    def path_to_folder(self):
        """ Gets path to checklist excel files folder. Creates folder is not exists """
        path = Path(os.path.join(settings.MEDIA_ROOT, 'checklist_excels'))

        if not path.exists():
            os.mkdir(os.path.join(settings.MEDIA_ROOT, 'checklist_excels'))

        return path

    @property
    def path_to_file(self):
        return os.path.join(self.path_to_folder, self.file_name)

    def _generate_name(self):
        """ Generates checklist file name """
        return 'Checklist_{}.xlsx'.format(self.checklist.id)

    def convert(self):
        """ Converts checklist to xls file """
        self._set_columns_width()
        self._generate_header()
        last_row = self._generate_statistics()
        self._generate_questions_block(last_row+2)
        return save_virtual_workbook(self.workbook)

    def _set_columns_width(self):
        """ Sets proper column width for sheet """
        columns_width_map = {'B': 70, 'C': 30, 'D': 15, 'E': 40, 'F': 40}

        for column_name, column_width in columns_width_map.items():
            self.ws.column_dimensions[column_name].width = column_width

    def _generate_header(self):
        """ Generates checklist file header """

        ws = self.workbook.active
        self.ws['B1'] = 'Дата проверки'
        self.ws['B1'].font = self.styles['bold_font']
        self.ws['B1'].border = self.styles['border_1']

        date_string = 'Не указано'
        if self.checklist.visit_date:
            date_string = self.checklist.visit_date.strftime('%d.%m.%Y')

        self.ws['C1'] = date_string
        self.ws['C1'].alignment = self.styles['centered_text']
        self.ws['C1'].border = self.styles['border_1']

        self.ws['B2'] = 'ДЦ'
        self.ws['B2'].font = self.styles['bold_font']
        self.ws['B2'].border = self.styles['border_1']

        if not self.checklist.check_obj is None:
            self.ws['C2'] = self.checklist.check_obj.target.title
        else:
            self.ws['C2'] = 'Не указано'
        self.ws['C2'].border = self.styles['border_1']

    def _generate_statistics(self):
        """ Generates checklist statistics """

        self.fill_statistics_header()

        current_row = 6
        for section_name, section_data in self.statistics.sections.items():
            self.fill_statistics_section_row(current_row, section_name, section_data)
            current_row += 1

            for subsection_name, subsection_data in section_data['subsections'].items():
                self.fill_statistics_section_row(current_row, subsection_name, subsection_data)
                current_row += 1

        return current_row

    def fill_statistics_header(self):
        """Fill statistics block header with data."""
        self.ws['C5'] = 'Балл'
        self.ws['C5'].font = self.styles['bold_font']
        self.ws['C5'].alignment = self.styles['centered_text']
        self.ws['C5'].border = self.styles['border_1']

        self.ws['D5'] = 'Макс.'
        self.ws['D5'].font = self.styles['bold_font']
        self.ws['D5'].alignment = self.styles['centered_text']
        self.ws['D5'].border = self.styles['border_1']

        self.ws['E5'] = '%'
        self.ws['E5'].font = self.styles['bold_font']
        self.ws['E5'].alignment = self.styles['centered_text']
        self.ws['E5'].border = self.styles['border_1']

    def fill_statistics_section_row(self, current_row, section_name, section_data):
        """Fill data for one statistics block row(section data)."""
        # section name
        self.ws['B{}'.format(current_row)] = section_name
        self.ws['B{}'.format(current_row)].alignment = self.styles['centered_text']
        self.ws['B{}'.format(current_row)].border = self.styles['border_1']

        # section max points available
        self.ws['C{}'.format(current_row)] = section_data['points']
        self.ws['C{}'.format(current_row)].alignment = self.styles['centered_text']
        self.ws['C{}'.format(current_row)].border = self.styles['border_1']

        # section total got points
        self.ws['D{}'.format(current_row)] = section_data['points_total']
        self.ws['D{}'.format(current_row)].alignment = self.styles['centered_text']
        self.ws['D{}'.format(current_row)].border = self.styles['border_1']

        # section completion percentage
        self.ws['E{}'.format(current_row)] = '{}%'.format(section_data['percentage'])
        self.ws['E{}'.format(current_row)].alignment = self.styles['centered_text']
        self.ws['E{}'.format(current_row)].border = self.styles['border_1']

    def _generate_questions_block(self, start_row):
        """Fill data for question block."""
        self._generate_questions_block_header(start_row)
        current_row = start_row

        for section in self.checklist.sections.filter(parent=None):
            current_row += 1
            self._generate_section_header(section, current_row, self.statistics)

            for question in section.questions.all():
                current_row += 1
                self._fill_question_data(question, current_row)

            for subsection in section.subsections.all():
                current_row += 1
                statistics = self.statistics.sections.get(subsection.name.value, '')
                self._generate_subsection_header(subsection, current_row, statistics)

                for question in subsection.questions.all():
                    current_row += 1
                    self._fill_question_data(question, current_row)

        self._end_table(current_row)

    def _generate_questions_block_header(self, row):
        """Fill questions block header"""
        self.ws['B{}'.format(row)] = 'Проявления в работе продавца-консультанта'
        self.ws['B{}'.format(row)].font = self.styles['bold_font']
        self.ws['B{}'.format(row)].alignment = self.styles['centered_text']
        self.ws['B{}'.format(row)].border = self.styles['border_2']

        self.ws['C{}'.format(row)] = 'Критерии оценки'
        self.ws['C{}'.format(row)].font = self.styles['bold_font']
        self.ws['C{}'.format(row)].alignment = self.styles['centered_text']
        self.ws['C{}'.format(row)].border = self.styles['border_2']

        self.ws['D{}'.format(row)] = 'Выполнение'
        self.ws['D{}'.format(row)].font = self.styles['bold_font']
        self.ws['D{}'.format(row)].alignment = self.styles['centered_text']
        self.ws['D{}'.format(row)].border = self.styles['border_2']

        self.ws['E{}'.format(row)] = 'Комментарий клиента'
        self.ws['E{}'.format(row)].font = self.styles['bold_font']
        self.ws['E{}'.format(row)].alignment = self.styles['centered_text']
        self.ws['E{}'.format(row)].border = self.styles['border_2']

        self.ws['F{}'.format(row)] = 'Апелляция'
        self.ws['F{}'.format(row)].font = self.styles['bold_font']
        self.ws['F{}'.format(row)].alignment = self.styles['centered_text']
        self.ws['F{}'.format(row)].border = self.styles['border_2']

        # make row height
        self.ws.row_dimensions[row].height = 30

    def _generate_section_header(self, section, row, statistics):
        """Generate questions block section header"""

        # section name
        self.ws['B{}'.format(row)] = section.name.value
        self.ws['B{}'.format(row)].font = self.styles['bold_font']
        self.ws['B{}'.format(row)].border = self.styles['border_2']
        self.ws['B{}'.format(row)].alignment = self.styles['centered_text']
        self.ws['B{}'.format(row)].fill = self.styles['bg_grey']

        # section ma[ points available
        self.ws['C{}'.format(row)] = statistics.sections.get(section.name.value, '')['points_total']
        self.ws['C{}'.format(row)].font = self.styles['bold_font']
        self.ws['C{}'.format(row)].alignment = self.styles['centered_text']
        self.ws['C{}'.format(row)].border = self.styles['border_2']
        self.ws['C{}'.format(row)].fill = self.styles['bg_grey']

        # section points got
        self.ws['D{}'.format(row)] = statistics.sections.get(section.name.value, '')['points']
        self.ws['D{}'.format(row)].font = self.styles['bold_font']
        self.ws['D{}'.format(row)].alignment = self.styles['centered_text']
        self.ws['D{}'.format(row)].border = self.styles['border_2']
        self.ws['D{}'.format(row)].fill = self.styles['bg_grey']

        # fill other cell on a row with color
        self.ws['E{}'.format(row)].fill = self.styles['bg_grey']
        self.ws['E{}'.format(row)].border = self.styles['border_2']
        self.ws['F{}'.format(row)].fill = self.styles['bg_grey']
        self.ws['F{}'.format(row)].border = self.styles['border_2']

        # make row height
        self.ws.row_dimensions[row].height = 25

    def _generate_subsection_header(self, subsection, row, statistics):
        """Generate questions block subsection header."""
        self.ws['B{}'.format(row)] = subsection.name.value
        self.ws['B{}'.format(row)].font = self.styles['bold_font']
        self.ws['B{}'.format(row)].alignment = self.styles['centered_text']
        self.ws['B{}'.format(row)].border = self.styles['border_2']
        self.ws['B{}'.format(row)].fill = self.styles['bg_grey']

        self.ws['C{}'.format(row)] = statistics.get('points_total', 0)
        self.ws['C{}'.format(row)].font = self.styles['bold_font']
        self.ws['C{}'.format(row)].alignment = self.styles['centered_text']
        self.ws['C{}'.format(row)].border = self.styles['border_2']
        self.ws['C{}'.format(row)].fill = self.styles['bg_grey']

        self.ws['D{}'.format(row)] = statistics.get('points', 0)
        self.ws['D{}'.format(row)].font = self.styles['bold_font']
        self.ws['D{}'.format(row)].border = self.styles['border_2']
        self.ws['D{}'.format(row)].alignment = self.styles['centered_text']
        self.ws['D{}'.format(row)].fill = self.styles['bg_grey']

        # fill other cell on a row with color
        self.ws['E{}'.format(row)].fill = self.styles['bg_grey']
        self.ws['E{}'.format(row)].border = self.styles['border_2']
        self.ws['F{}'.format(row)].fill = self.styles['bg_grey']
        self.ws['F{}'.format(row)].border = self.styles['border_2']

        # make row height
        self.ws.row_dimensions[row].height = 25

    def _fill_question_data(self, question, current_row):
        """Fill data for given question."""
        # fill question text cell
        self.parser.feed(question.text)
        self.ws['B{}'.format(current_row)] = self.parser.data
        self.parser.clean()
        self.ws['B{}'.format(current_row)].font = self.styles['question_font']
        self.ws['B{}'.format(current_row)].border = self.styles['question_cell_border']
        self.ws['B{}'.format(current_row)].alignment = self.styles['question_align']

        # fill question answer options cell
        self.ws['C{}'.format(current_row)] = question.answer.get_answer_options()
        self.ws['C{}'.format(current_row)].alignment = self.styles['wrap_text']
        self.ws['C{}'.format(current_row)].border = self.styles['usual_border']

        # fill question answer cell
        self.ws['D{}'.format(current_row)] = question.answer.value
        self.ws['D{}'.format(current_row)].border = self.styles['usual_border']
        self.ws['D{}'.format(current_row)].alignment = self.styles['centered_text']

        # fill question perform comment cell
        if question.answer.performer_comment:
            self.ws['E{}'.format(current_row)] = question.answer.performer_comment
        self.ws['E{}'.format(current_row)].border = self.styles['usual_border']

        # fill question appeal cell
        if question.answer.appeal_comment:
            self.ws['F{}'.format(current_row)] = question.answer.appeal_comment
        self.ws['F{}'.format(current_row)].border = self.styles['usual_border_right']

        self.ws.row_dimensions[current_row].height = 40

    def _end_table(self, table_last_row):
        """Make proper borders for questions block last row."""
        self.ws[f'B{table_last_row}'].border = self.styles['last_row_left']
        self.ws[f'C{table_last_row}'].border = self.styles['last_row_center']
        self.ws[f'D{table_last_row}'].border = self.styles['last_row_center']
        self.ws[f'E{table_last_row}'].border = self.styles['last_row_center']
        self.ws[f'F{table_last_row}'].border = self.styles['last_row_right']
